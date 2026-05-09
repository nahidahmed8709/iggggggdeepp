import time
import random
import string
import pyotp
from faker import Faker
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import os

# ======= CONFIG =======
EXCEL_FILE = "instagram_accounts.xlsx"
# ======================

fake = Faker()

def random_password(length=12):
    """Generate a random password with at least one digit and one special char."""
    chars = string.ascii_letters + string.digits + "!@#$%*"
    while True:
        pwd = ''.join(random.choices(chars, k=length))
        if any(c.isdigit() for c in pwd) and any(c in "!@#$%*" for c in pwd):
            return pwd

def random_birthdate(min_age=18, max_age=30):
    """Return a random birthdate in the format 'MM DD YYYY' (three separate values)."""
    date = fake.date_of_birth(minimum_age=min_age, maximum_age=max_age)
    return date.strftime("%m"), date.strftime("%d"), date.strftime("%Y")

def save_to_excel(email, username, password, twofa_secret):
    """Append a new account row to the Excel file."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Accounts"
        ws.append(["Email", "Username", "Password", "2FA_Secret"])
        wb.save(EXCEL_FILE)

    wb = Workbook()
    # Loading existing workbook with openpyxl requires load_workbook, but simpler: always append to new file? We'll use a manual approach.
    # Actually, let's use a simpler method: keep data in list and save at end. But for each account, we want immediate save.
    # I'll modify to append to a CSV after each account, then at end convert to xlsx? Or just keep in memory and ask user to save all.
    # To match the request "saved in xlsx in file section", I'll save an xlsx with all accounts created in one session.
    # We'll store accounts in a global list and write once when user exits the loop.
    pass  # We'll handle saving in main loop with a list.

def create_instagram_account(email, otp):
    """Main routine: creates an Instagram account, enables 2FA, extracts secret."""
    # Random data
    full_name = fake.name()
    username = fake.user_name() + str(random.randint(10, 99))
    password = random_password()
    month, day, year = random_birthdate()

    # Chrome options for Termux (headless mode may not work; we need a display)
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # If you have Termux:X11 or VNC, you can run without headless. For automation, headless is better:
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    # Use Chromium binary installed by pkg
    chrome_options.binary_location = "/data/data/com.termux/files/usr/bin/chromium-browser"

    driver = webdriver.Chrome(options=chrome_options)
    driver.set_window_size(412, 915)  # mobile viewport to avoid layout issues

    try:
        # 1. Open Instagram signup page
        driver.get("https://www.instagram.com/accounts/emailsignup/")
        time.sleep(3)

        # 2. Fill email
        email_input = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.NAME, "emailOrPhone"))
        )
        email_input.send_keys(email)

        # 3. Fill full name
        name_input = driver.find_element(By.NAME, "fullName")
        name_input.send_keys(full_name)

        # 4. Fill username
        username_input = driver.find_element(By.NAME, "username")
        username_input.send_keys(username)

        # 5. Fill password
        password_input = driver.find_element(By.NAME, "password")
        password_input.send_keys(password)

        # 6. Click Sign Up
        signup_btn = driver.find_element(By.XPATH, "//button[@type='submit']")
        signup_btn.click()
        time.sleep(3)

        # 7. Enter birthday
        # Instagram asks for birthday on next screen (sometimes appears after signup)
        # Select month, day, year from dropdowns
        month_select = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//select[@title='Month:']"))
        )
        month_select.click()
        month_select.find_element(By.XPATH, f"//option[@value='{month}']").click()

        day_select = driver.find_element(By.XPATH, "//select[@title='Day:']")
        day_select.click()
        day_select.find_element(By.XPATH, f"//option[@value='{day}']").click()

        year_select = driver.find_element(By.XPATH, "//select[@title='Year:']")
        year_select.click()
        year_select.find_element(By.XPATH, f"//option[@value='{year}']").click()

        # Click Next
        next_btn = driver.find_element(By.XPATH, "//button[text()='Next']")
        next_btn.click()
        time.sleep(2)

        # 8. Enter confirmation code (OTP from email)
        print(f"[!] Check email {email} for the confirmation code.")
        # The user already gave us the OTP as input, we can use it directly
        code_fields = WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.XPATH, "//input[@aria-label='Confirmation Code']"))
        )
        if len(code_fields) == 6:
            for i, digit in enumerate(otp):
                code_fields[i].send_keys(digit)
        else:
            # Sometimes it's a single field
            code_input = driver.find_element(By.XPATH, "//input[@aria-label='Confirmation Code']")
            code_input.send_keys(otp)

        # Click Next/Confirm
        confirm_btn = driver.find_element(By.XPATH, "//button[text()='Next']")
        confirm_btn.click()
        time.sleep(5)

        # 9. Skip "Save Login Info" if shown
        try:
            not_now = driver.find_element(By.XPATH, "//button[text()='Not Now']")
            not_now.click()
            time.sleep(2)
        except:
            pass

        # 10. Navigate to 2FA settings
        driver.get("https://www.instagram.com/accounts/two_factor_authentication/")
        time.sleep(3)

        # Click "Get Started" (if first time)
        try:
            get_started = driver.find_element(By.XPATH, "//button[contains(text(),'Get Started')]")
            get_started.click()
            time.sleep(2)
        except:
            pass

        # Choose "Authentication App"
        auth_app_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[text()='Authentication App']/.."))
        )
        auth_app_btn.click()
        time.sleep(2)

        # Click "Next" on the introductory screen
        next_btn_2fa = driver.find_element(By.XPATH, "//button[text()='Next']")
        next_btn_2fa.click()
        time.sleep(2)

        # Extract the 2FA secret key (it's shown in a text field)
        secret_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "two_factor_secret_key"))
        )
        twofa_secret = secret_input.get_attribute("value")

        # Click "Done" (or "Next") to finish setup
        done_btn = driver.find_element(By.XPATH, "//button[text()='Done']")
        done_btn.click()
        time.sleep(2)

        print(f"✅ Account created: {username}")
        print(f"   2FA Secret: {twofa_secret}")

        driver.quit()
        return {
            "email": email,
            "username": username,
            "password": password,
            "2fa_secret": twofa_secret
        }

    except Exception as e:
        print(f"❌ Error: {e}")
        driver.quit()
        return None

def main():
    accounts = []

    while True:
        email = input("Enter email (or 'done' to finish): ").strip()
        if email.lower() == 'done':
            break
        if not email:
            print("Email cannot be empty.")
            continue

        otp = input("Enter the OTP code received: ").strip()
        if not otp:
            print("OTP required.")
            continue

        print("\n[+] Creating Instagram account...")
        acc_data = create_instagram_account(email, otp)
        if acc_data:
            accounts.append(acc_data)
        else:
            print("Account creation failed. Skipping.")

    # Save all accounts to Excel
    if accounts:
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Accounts"
        ws.append(["Email", "Username", "Password", "2FA_Secret"])
        for acc in accounts:
            ws.append([acc["email"], acc["username"], acc["password"], acc["2fa_secret"]])
        wb.save(EXCEL_FILE)
        print(f"\n📁 Saved {len(accounts)} account(s) to '{EXCEL_FILE}'")
    else:
        print("No accounts created.")

if __name__ == "__main__":
    main()